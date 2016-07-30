#!/usr/bin/env python

from openpyxl import load_workbook
import re
import json
import sys
import os

YEAR = '2013'
if len(sys.argv) != 2:
    print("Usage: census.py <dataset>\n (where dataset can be found in input-datasets)")
    exit(1)
NAME = sys.argv[1]

wb = load_workbook(
    filename='input-datasets/%s.xlsx' % NAME,
    read_only=True
)

worksheet_matcher= re.compile(r'area unit', re.I)
ws = None
for name in wb.sheetnames:
    if worksheet_matcher.search(name):
        ws = wb[name]
        break
if ws is None:
    raise Exception("Failed to detect worksheet")

rows = iter(ws.rows)

# skip crap at top of table
for _ in range(0, 8):
    next(rows)

# parse headers
headers = {}

area_unit_column = None
area_unit_matcher = re.compile(r'area unit code', re.I)

census_year_matcher = re.compile(r'[\d]+ Census', re.I)

h1 = next(rows)
h2 = next(rows)
header_prefix = h1[0].value
for i in range(0, len(h2)):
    # read the headers for this column
    header_suffix = h2[i].value
    next_prefix = h1[i].value

    # handle empty (merged) cells
    if next_prefix is not None:
        header_prefix = next_prefix
    elif header_suffix is None:
        break

    # skip old data
    if census_year_matcher.match(header_prefix) and not header_prefix.startswith(YEAR):
        continue

    # headers generally have categories, we need those
    if header_suffix is None:
        header = [ header_prefix ]
    elif header_prefix is not header_suffix:
        header = [ header_prefix, header_suffix ]
    else:
        header = [ header_suffix ]

    headers[i] = header

    if area_unit_matcher.match(header[0]):
        if area_unit_column is not None:
            raise Exception("Detected multiple conflicting area unit columns!")
        area_unit_column = i

if area_unit_column is None:
    raise Exception("Failed to detect area unit column!")

output = 'processed-datasets/%s' % NAME
os.makedirs(output, exist_ok=True)
os.chdir(output)
for row in rows:
    area_unit = row[area_unit_column].value
    # Skip empty rows (should just be the final/total row at the end)
    if area_unit is None:
        continue
    obj = {}
    # Create the data structure from the row
    for column in headers.keys():
        pos = obj
        last_key = None
        for key in headers[column]:
            if last_key is not None:
                pos = pos[last_key]
            if key not in pos:
                pos[key] = {}
            last_key = key
        pos[last_key] = row[column].value
    # Write the data to a file
    fh = open(area_unit, 'w')
    try:
        json.dump(obj, fh)
    finally:
        fh.close()
